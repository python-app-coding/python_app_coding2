# coding = utf8
import os.path

import openpyxl
import pandas as pd
import numpy as np


def make_demo_file(file):
    from sec113_write_sheets import make_faker_dataframe
    df1, df2 = make_faker_dataframe(10)
    df2.to_excel(file)


def demo_openpyxl(file="temp_demo_openpyxl.xlsx"):
    wk = openpyxl.load_workbook(file)
    sh = wk.active

    # get cell
    for row in sh.rows:
        for cell in row:
            cell_value = cell.value
            print(cell_value, end=', ')
        print('')

    # set cell
    for i, row in enumerate(sh.rows):
        for j, cell in enumerate(row):
            if i > 0 and j == 3:
                cell.value = 100

    # get data1
    for row in sh.rows:
        for cell in row:
            print(cell.value, end=', ')
        print('')

    # get cell 2
    for i in range(sh.max_row):
        for j in range(sh.max_column):
            cell_value = sh.cell(row=i+1, column=j+1).value     # worksheet.cell的行列序号从1开始
            print(cell_value, end=', ')
        print('')

    # set cell 2
    for i in range(sh.max_row):
        for j in range(sh.max_column):
            if i > 0 and j == 4:
                sh.cell(row=i+1, column=j+1, value=200)         # 使用worksheet.cell直接赋值
            cell_value = sh.cell(row=i+1, column=j+1).value
            print(cell_value, end=', ')
        print('')

    wk.close()


if __name__ == '__main__':
    demo_file = 'temp_demo_openpyxl.xlsx'
    if not os.path.isfile(demo_file):
        make_demo_file(demo_file)
    demo_openpyxl(demo_file)
