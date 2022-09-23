# coding = utf8

import openpyxl
import pandas as pd
import numpy as np


def demo_fill_data_to_excel():

    excel_file0 = 'temp_demo_fill_data_0.xlsx'
    excel_file1 = 'temp_demo_fill_data_1.xlsx'
    sheet_name = 'score2020'

    df = pd.DataFrame(
        data={
            '姓名': ['张飞飞', '王康泰', '李宋儒', '赵统成', '杨德伟', '孙丽翠'],
            '语文': [None for _ in range(6)],
            '数学': [None for _ in range(6)],
            '外语': [None for _ in range(6)],
            '总分': [None for _ in range(6)]
        },
        index=['010' + str(j + 1) for j in range(6)]
    )
    df.index.name = '学号'

    # 各考生成绩数据
    data = np.array([[80, 90, 88],
                     [91, 92, 93],
                     [81, 90, 86],
                     [71, 82, 80],
                     [66, 72, 77],
                     [75, 52, 83]])

    # write DataFrame to Excel
    # 生成ExcelWriter对象，使用引擎openpyxl
    with pd.ExcelWriter(excel_file0, engine='openpyxl') as writer:
        df.to_excel(writer,
                    sheet_name=sheet_name,
                    index=True,
                    index_label='学号',
                    startrow=0
                    )

    # add title to excel_file0
    book = openpyxl.load_workbook(filename=excel_file0)
    worksheet = book.worksheets[0]
    # -- 在第一行前面插入一行， 准备给Excel表添加总标题
    worksheet.insert_rows(1)
    for inx, row in enumerate(worksheet.rows):
        if inx == 0:
            row[2].value = '2020年第一学期三班学生成绩报告'
    book.save(filename=excel_file0)
    book.close()

    df1 = pd.read_excel(excel_file0,
                        header=0,
                        names=['sno', 'name', 'yw', 'sx', 'wy', 'zf'],
                        dtype={'sno': str})

    # add title and score to excel_file from excel_file0
    book = openpyxl.load_workbook(filename=excel_file0)
    worksheet = book.worksheets[0]
    # worksheet.insert_rows(1)
    for inx, row in enumerate(worksheet.rows):
        if inx > 1:
            for j in range(2, 5):
                row[j].value = data[inx - 2, j - 2]
            row[5].value = sum(data[inx - 2, :])
    book.save(filename=excel_file1)
    book.close()

    df2 = pd.read_excel(io=excel_file1,
                        sheet_name=sheet_name,
                        header=1,
                        names=['sno', 'name', 'yw', 'sx', 'wy', 'zf'],
                        dtype={'sno': str},
                        )

    # 用于显示初始Excel表中数据情况
    kms = ['学号', '姓名', '语文', '数学', '外语', '总分']
    excel_str = '学号' + ''.join(['{:>10s}'.format(s) for s in kms[1:]]) + '\n'
    for ind, row in df.iterrows():
        excel_str += str(ind)
        for col in row:
            excel_str += '{:>10s}'.format(str(col) if col else '---')
        excel_str += '\n'

    print(
        f"# Excel文件demo.xlsx:\n"
        f"     2020年第一学期三班学生成绩报告\n"
        f"{excel_str}\n\n"

        f"# 读取Excel文件\n"
        f" >>> df1 = pd.read_excel(io='demo_excel0.xlsx',\n"
        f" ...                    header=1,\n"
        f" ...                    names=['sno', 'name', 'yw', 'sx', 'wy', 'zf'], \n"
        f" ...                    dtype={{'sno': str}})\n"
        f" >>> df1\n"
        f"{df1}\n"

        f"# 学生成绩数据\n"
        f" >>> data\n"
        f"array(\n{data})\n"

        """
        # 使用openpyxl模块载入工作簿，写入成绩数据并保存Excel文件
         >>> import openpyxl
         >>> book = openpyxl.load_workbook(filename='demo.xlsx')
         # 第一个工作表数据
         >>> worksheet = book.worksheets[0]             
         >>> for idx, row in enumerate(worksheet.rows):
         ...     if idx < 2:
         ...        continue
         ...     # 从第二行开始添加成绩数据, 列分别为：col=2(yw), 3(sx), 4(wy), 5(zf)
         ...     for j in range(2, 5):
         ...         row[j].value = data[idx-2，j-2]     # 填入各科成绩 
         ...     row[5].value = sum(data[idx-2])         # 计算总分
         # 保存数据到Excel文件
         >>> book.save(filename=‘demo_excel.xlsx’)
        """
        f"\n"
        
        f"# 读取写入成绩后的Excel文件\n"
        f" >>> df2 = pd.read_excel(io='demo_excel.xlsx',\n"
        f" ...                     sheet_name='demo_excel1.xlsx')\n"
        f" ...                     header=1,\n"
        f" ...                     names=['sno', 'name', 'yw', 'sx', 'wy', 'zf'],\n"
        f" ...                     dtype={{'sno': str}}\n"
        f" ...                     )\n"
        f" >>> df2\n"
        f"{df2}"
    )


if __name__ == '__main__':
    demo_fill_data_to_excel()
